from flask import Flask, request, jsonify, render_template, send_file
import pytesseract
from pytesseract import Output
import cv2
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
import logging
import time

# ✅ 네 PC 기준 Tesseract 경로
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

app = Flask(__name__)
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')


@app.route('/')
def home():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    uploads_folder = 'uploads'
    outputs_folder = 'outputs'
    os.makedirs(uploads_folder, exist_ok=True)
    os.makedirs(outputs_folder, exist_ok=True)

    # 기존 업로드 파일 삭제
    for f in os.listdir(uploads_folder):
        os.remove(os.path.join(uploads_folder, f))

    # 파일 저장
    unique_filename = f"{int(time.time())}_{file.filename}"
    filepath = os.path.join(uploads_folder, unique_filename)
    file.save(filepath)

    logging.debug(f"Processing file: {filepath}")

    try:
        # =======================
        # 이미지 전처리
        # =======================
        image = cv2.imread(filepath)
        if image is None:
            return jsonify({'error': 'Failed to read image file'}), 400

        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        gray = cv2.convertScaleAbs(gray, alpha=1.5, beta=0)
        gray = cv2.fastNlMeansDenoising(gray, None, 10, 7, 21)
        binary = cv2.threshold(gray, 0, 255,
                               cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]

        # =======================
        # OCR (텍스트 + 위치)
        # =======================
        custom_config = r'--psm 6 --oem 3'
        ocr_data = pytesseract.image_to_data(
            binary, lang='kor+eng', config=custom_config, output_type=Output.DICT
        )

        full_text = pytesseract.image_to_string(
            binary, lang='kor+eng', config=custom_config
        )
        logging.debug(f"=== FULL TEXT ===\n{full_text}\n================")

        # =======================
        # 엑셀 워크북 생성
        # =======================
        wb = Workbook()
        ws = wb.active
        ws.title = "Table OCR"

        n_boxes = len(ocr_data['text'])
        line_dict = {}

        # 줄 단위 그룹핑
        for i in range(n_boxes):
            text = ocr_data['text'][i].strip()
            conf_str = ocr_data['conf'][i]

            # 신뢰도 파싱
            try:
                conf = int(float(conf_str))
            except:
                conf = -1

            if conf > 30 and text:
                top = ocr_data['top'][i]
                left = ocr_data['left'][i]

                # 👉 여기 값(15)이 줄 그룹핑 민감도
                line_key = round(top / 15)

                if line_key not in line_dict:
                    line_dict[line_key] = []

                line_dict[line_key].append({
                    'left': left,
                    'text': text
                })

        # 세로(위→아래) 정렬
        sorted_lines = sorted(line_dict.items())
        excel_row = 1

        for _, words in sorted_lines:
            # 가로(왼→오) 정렬
            words.sort(key=lambda x: x['left'])

            line_text = ' '.join([w['text'] for w in words])

            # 👉 여기 값(100)이 컬럼 나누기 기준
            column_positions = []
            prev_left = 0

            for w in words:
                if w['left'] - prev_left > 100 and len(column_positions) > 0:
                    column_positions.append(w.copy())
                elif len(column_positions) == 0:
                    column_positions.append(w.copy())
                else:
                    column_positions[-1]['text'] += ' ' + w['text']
                prev_left = w['left'] + len(w['text']) * 10

            # =======================
            # 엑셀에 쓰기
            # =======================
            if len(column_positions) > 1:
                # 다중 컬럼 행
                for col_idx, col_data in enumerate(column_positions, start=1):
                    cell = ws.cell(row=excel_row, column=col_idx, value=col_data['text'])
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell.font = Font(name='Arial', size=11)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            else:
                # 한 컬럼 행 (key:value 패턴 처리)
                if ':' in line_text and not line_text.startswith(':'):
                    parts = line_text.split(':', 1)

                    key_cell = ws.cell(row=excel_row, column=1, value=parts[0].strip())
                    key_cell.font = Font(name='Arial', size=11, bold=True)
                    key_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

                    if len(parts) > 1:
                        value_cell = ws.cell(row=excel_row, column=2, value=parts[1].strip())
                        value_cell.font = Font(name='Arial', size=11)

                    for col in [1, 2]:
                        ws.cell(row=excel_row, column=col).border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        ws.cell(row=excel_row, column=col).alignment = Alignment(
                            horizontal='left', vertical='top', wrap_text=True
                        )
                else:
                    cell = ws.cell(row=excel_row, column=1, value=line_text)
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell.font = Font(name='Arial', size=11)

                    if line_text.isupper() or len(line_text) < 30:
                        cell.font = Font(name='Arial', size=11, bold=True)
                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

            excel_row += 1

        # 컬럼 너비
        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 25
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50

        # 엑셀 저장
        output_filename = f"{os.path.splitext(unique_filename)[0]}.xlsx"
        output_path = os.path.join(outputs_folder, output_filename)
        wb.save(output_path)

        logging.debug(f"Excel file saved to: {output_path}")

        return jsonify({
            'message': 'File processed successfully',
            'download_url': f"/download/{output_filename}",
            'rows_created': excel_row - 1
        }), 200

    except Exception as e:
        logging.error(f"Error processing file {filepath}: {e}")
        import traceback
        logging.error(traceback.format_exc())
        return jsonify({'error': f"An error occurred: {str(e)}"}), 500


@app.route('/download/<filename>')
def download_file(filename):
    filepath = os.path.join('outputs', filename)
    if not os.path.exists(filepath):
        return "File not found", 404
    return send_file(filepath, as_attachment=True)


if __name__ == '__main__':
    os.makedirs('uploads', exist_ok=True)
    os.makedirs('outputs', exist_ok=True)
    app.run(debug=True)
